import os
from fileinput import filename

import openpyxl
import xlrd
import pandas as pd
from openpyxl import Workbook

"""Reading multiple files .xlsx, count and print"""


path = r"X:\55838\55838-07\06 - Pipe Support Engineering\6.03 - Non Advance Steel CAD Drawings\55838-07 support dwgs\BOM"
bom_dir = os.listdir(r"X:\55838\55838-07\06 - Pipe Support Engineering\6.03 - Non Advance "
                     r"Steel CAD Drawings\55838-07 support dwgs\BOM")


item_list = {}
def read_files(file_name):
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb["BOM"]

    multiply_bom = 1
    print(file_name)

    for i in ws['F1':'F1']:
        multiply_bom = i[0].value

    for i in ws['A3':'H500']:
        if i[0].value:

            qty = int(i[1].value)
            part_number_grade = str(i[2].value)
            description = str(i[3].value)

            item_weight = 0
            total_weight = 0
            try:
                item_weight = float(i[5].value)
            except:
                item_weight = 0

            total_weight = float(i[6].value)

            unique_key = (description + "&" + part_number_grade).strip().replace(" ", "")

            if unique_key not in item_list.keys():
                item_list[unique_key] = [description, qty, part_number_grade, item_weight, total_weight]
            else:
                item_list[unique_key][1] += qty * int(multiply_bom)
                item_list[unique_key][4] += total_weight * int(multiply_bom)

            # if description in item_list.keys():
            #     if item_list[description][1] == part_number_grade:
            #         item_list[description][0] += qty * int(multiply_bom)
            #         item_list[description][2] = item_weight * int(multiply_bom)
            #         item_list[description][3] += total_weight * int(multiply_bom)
            #     else:
            #         try:
            #             if item_list[description + " "][1] == part_number_grade:
            #                 item_list[description][0] += qty * int(multiply_bom)
            #                 item_list[description][2] = item_weight * int(multiply_bom)
            #                 item_list[description][3] += total_weight * int(multiply_bom)
            #         except:
            #             item_list[description + " "] = [qty, part_number_grade, item_weight, total_weight]

            # else:
            #     item_list[description] = [qty, part_number_grade, item_weight, total_weight]
    # print(file_name, " - done\n")

def just_print(filename):
    print(filename)
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb["BOM"]

    multiply_bom = 1

    for i in ws['A3':'H500']:
        if i[0].value:
            qty = int(i[1].value)
            part_number_grade = str(i[2].value)
            description = str(i[3].value)

            item_weight = 0
            total_weight = 0
            try:
                item_weight = float(i[5].value)
            except:
                item_weight = 0

            total_weight = float(i[6].value)
            print(f"{description}&{qty}&{part_number_grade}&{total_weight}")



for file_name in bom_dir:
    if "~" not in file_name:
        file_name = path + "\\" + file_name
        read_files(file_name)

print("Description& QTY& Part number/grade & Item weight & Total weight")
for key in item_list:
    print(item_list[key][0], "&", item_list[key][1], "&", item_list[key][2], "&", item_list[key][3], "&", item_list[key][4])





