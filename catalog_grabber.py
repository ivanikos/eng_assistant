from fileinput import filename

catalogs_folder = r"C:\Users\ivaign\OneDrive - United Conveyor Corp\Desktop\P3DAd\Catalogs summary"

# Example usage
# file_path = catalogs_folder + f"\\Ucarbstl.xlsx"  # Replace with your Excel file path


column_names = ['Sizes', 'Short Description', 'Long Description (Family)', 'UCCPARTNO', 'UCCDWGNO']  # Replace with the names of the columns you need

import openpyxl as xl
import os
import xlsxwriter


summary_list = [["Size", "Short Desc.", "Long Desc.", "UCC_PN", "UCC_DWG_NO", "Catalog"]]

def grab_inf(file_path,catalog_name):
    # global summary_list
    wb_catalog = xl.load_workbook(file_path)
    for sheet in wb_catalog.sheetnames:
        work_sheet = wb_catalog[f"{sheet}"]
        # print(sheet, work_sheet)
        sizes = 0
        s_description = 0
        l_description = 0
        ucc_pn = 0
        ucc_dwgno = 0
        for i in work_sheet["A2":"CI2"]:
            for j in range(0, (len(i) - 1)):
                if i[j].value == "Sizes":
                    sizes = j
                if i[j].value == "Short Description":
                    s_description = j
                if i[j].value == "Long Description (Family)":
                    l_description = j
                if i[j].value == "UCCPARTNO":
                    ucc_pn = j
                if i[j].value == "UCCDWGNO":
                    ucc_dwgno = j

        for i in work_sheet["A3":"CI100"]:
            if i[sizes].value:
                summary_list.append([i[sizes].value, i[s_description].value, i[l_description].value, i[ucc_pn].value, i[ucc_dwgno].value, catalog_name])
                # print([i[sizes].value, i[s_description].value, i[l_description].value, i[ucc_pn].value, i[ucc_dwgno].value])


for catalog_name in os.listdir(catalogs_folder):
    if ".xlsx" in catalog_name:
        file_path = catalogs_folder + f"\\{catalog_name}"
        grab_inf(file_path, catalog_name)
        print(catalog_name, " - checked!")

for i in summary_list:
    print(i)


workbook_summary = xlsxwriter.Workbook(f'{catalogs_folder}\\Catalogs_summary.xlsx')
ws = workbook_summary.add_worksheet("summary")

for i, (one, two, three, four, five,six) in enumerate(summary_list):
    ws.write(f"A{i}", one)
    ws.write(f"B{i}", two)
    ws.write(f"C{i}", three)
    ws.write(f"D{i}", four)
    ws.write(f"E{i}", five)
    ws.write(f"F{i}", six)

workbook_summary.close()
print("success!")



